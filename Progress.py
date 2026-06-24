from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.differential import DifferentialStyle

wb = Workbook()

# ── Color palette ──────────────────────────────────────────────────────────────
NAVY      = "0A2D5E"
BLUE      = "0D4EA6"
GOLD      = "F5C518"
LIGHT_BG  = "F0F4F8"
WHITE     = "FFFFFF"
GREEN_BG  = "EAF3DE"
GREEN_FG  = "3B6D11"
AMBER_BG  = "FAEEDA"
AMBER_FG  = "854F0B"
RED_BG    = "FCEBEB"
RED_FG    = "A32D2D"
BLUE_BG   = "E6F1FB"
BLUE_FG   = "185FA5"
PURPLE_BG = "EEEDFE"
PURPLE_FG = "534AB7"
TEAL_BG   = "E1F5EE"
TEAL_FG   = "0F6E56"
GRAY_BG   = "F1EFE8"
GRAY_FG   = "5F5E5A"
ROW_ALT   = "F8FAFC"
BORDER_C  = "D1D5DB"

def side(c=BORDER_C, s="thin"):
    return Side(border_style=s, color=c)

def border(all=True, c=BORDER_C):
    s = side(c)
    if all:
        return Border(left=s, right=s, top=s, bottom=s)
    return Border(bottom=s)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=11, color="000000", name="Arial"):
    return Font(bold=bold, size=size, color=color, name=name)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_cell(ws, row, col, value, fnt=None, fil=None, aln=None, brd=None):
    c = ws.cell(row=row, column=col, value=value)
    if fnt: c.font = fnt
    if fil: c.fill = fil
    if aln: c.alignment = aln
    if brd: c.border = brd
    return c


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — COVER / DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
ws_cover = wb.active
ws_cover.title = "📊 Dashboard"
ws_cover.sheet_view.showGridLines = False
ws_cover.row_dimensions[1].height = 8
ws_cover.row_dimensions[2].height = 60
ws_cover.row_dimensions[3].height = 30
ws_cover.row_dimensions[4].height = 20
ws_cover.row_dimensions[5].height = 14
ws_cover.row_dimensions[6].height = 44
ws_cover.row_dimensions[7].height = 14
ws_cover.column_dimensions["A"].width = 3
for col in ["B","C","D","E","F","G","H"]:
    ws_cover.column_dimensions[col].width = 18

# Header banner
ws_cover.merge_cells("B2:H2")
c = ws_cover["B2"]
c.value = "RISET DATA KNOWLEDGE BASE — UNIVERSITAS MUHAMMADIYAH BANDUNG"
c.font = Font(bold=True, size=16, color=WHITE, name="Arial")
c.fill = fill(NAVY)
c.alignment = align("left", "center")
c.border = border()

ws_cover.merge_cells("B3:H3")
c = ws_cover["B3"]
c.value = "Checklist Kelengkapan Data untuk Chatbot Virtual Assistant UMBandung"
c.font = Font(bold=False, size=11, color=GOLD, name="Arial")
c.fill = fill(BLUE)
c.alignment = align("left", "center")

ws_cover.merge_cells("B4:H4")
c = ws_cover["B4"]
c.value = "Tidak ada data sensitif mahasiswa — hanya informasi umum yang dapat dipublikasikan"
c.font = Font(italic=True, size=9, color=GRAY_FG, name="Arial")
c.fill = fill(GRAY_BG)
c.alignment = align("left", "center")

# Stat boxes
stats = [
    ("Total Item\nRiset", "=COUNTA('✅ Checklist'!A:A)-1", BLUE_BG, BLUE_FG),
    ("Sudah\nSelesai",    "=COUNTIF('✅ Checklist'!E:E,\"✅ Selesai\")", GREEN_BG, GREEN_FG),
    ("Sedang\nDikerjakan","=COUNTIF('✅ Checklist'!E:E,\"🔄 Proses\")", AMBER_BG, AMBER_FG),
    ("Belum\nDimulai",    "=COUNTIF('✅ Checklist'!E:E,\"⏳ Belum\")", RED_BG, RED_FG),
    ("% Progress",        '=IFERROR(COUNTIF(\'✅ Checklist\'!E:E,"✅ Selesai")/(COUNTA(\'✅ Checklist\'!A:A)-1),0)', PURPLE_BG, PURPLE_FG),
]
stat_cols = ["B","C","D","E","F"]
for i, (label, formula, bg, fg) in enumerate(stats):
    col = stat_cols[i]
    ws_cover.row_dimensions[6].height = 50
    c_lbl = ws_cover[f"{col}6"]
    c_lbl.value = label
    c_lbl.font = Font(bold=False, size=9, color=fg, name="Arial")
    c_lbl.fill = fill(bg)
    c_lbl.alignment = align("center", "top", wrap=True)
    c_lbl.border = border()

    ws_cover.row_dimensions[7].height = 8

# Progress bar label row
ws_cover.row_dimensions[8].height = 20
ws_cover.merge_cells("B8:F8")
c = ws_cover["B8"]
c.value = "↑  Klik tab '✅ Checklist' untuk mulai mengisi status riset"
c.font = Font(italic=True, size=9, color=BLUE_FG, name="Arial")
c.alignment = align("left", "center")

# Instructions
ws_cover.row_dimensions[10].height = 18
ws_cover.merge_cells("B10:H10")
c = ws_cover["B10"]
c.value = "CARA PENGGUNAAN"
c.font = Font(bold=True, size=11, color=NAVY, name="Arial")
c.alignment = align("left", "center")

instructions = [
    ("1", "Buka tab '✅ Checklist'", "Berisi semua 60 item data yang perlu kamu riset, dikelompokkan per kategori."),
    ("2", "Isi kolom 'Status'", "Pilih dari dropdown: ✅ Selesai / 🔄 Proses / ⏳ Belum / ⛔ Skip"),
    ("3", "Isi kolom 'Sumber Data'", "Tulis URL atau nama sumber (misal: umbandung.ac.id/pmb, Instagram @umbandung)"),
    ("4", "Isi kolom 'Catatan / Data Ditemukan'", "Tulis ringkasan data yang kamu temukan — ini yang nantinya masuk ke knowledge base"),
    ("5", "Buka tab '📝 Template KB'", "Template siap pakai untuk menulis knowledge base dalam format yang bisa langsung dipakai di database_ID.py"),
]

for i, (num, title, desc) in enumerate(instructions):
    r = 11 + i
    ws_cover.row_dimensions[r].height = 28
    c_num = ws_cover.cell(row=r, column=2, value=num)
    c_num.font = Font(bold=True, size=10, color=WHITE, name="Arial")
    c_num.fill = fill(BLUE)
    c_num.alignment = align("center", "center")
    c_num.border = border()

    c_title = ws_cover.cell(row=r, column=3, value=title)
    c_title.font = Font(bold=True, size=10, color=NAVY, name="Arial")
    c_title.fill = fill(LIGHT_BG)
    c_title.alignment = align("left", "center")
    c_title.border = border()

    ws_cover.merge_cells(f"D{r}:H{r}")
    c_desc = ws_cover.cell(row=r, column=4, value=desc)
    c_desc.font = Font(size=9, color=GRAY_FG, name="Arial")
    c_desc.fill = fill(WHITE)
    c_desc.alignment = align("left", "center", wrap=True)
    c_desc.border = border()

# Sources section
ws_cover.row_dimensions[17].height = 18
ws_cover.merge_cells("B17:H17")
c = ws_cover["B17"]
c.value = "SUMBER RISET YANG DISARANKAN"
c.font = Font(bold=True, size=11, color=NAVY, name="Arial")
c.alignment = align("left", "center")

sources = [
    ("umbandung.ac.id",               "Website resmi — profil, visi misi, prodi, fasilitas, berita"),
    ("pmb.umbandung.ac.id",           "Portal PMB — jalur masuk, jadwal, syarat, biaya pendaftaran"),
    ("Instagram @umbandung",          "Info terkini, pengumuman PMB, kegiatan kampus"),
    ("pddikti.kemdikbud.go.id",       "Data akreditasi institusi dan per-program studi"),
    ("YouTube @umbandung",            "Video profil kampus, kegiatan, dan orientasi mahasiswa"),
    ("Telepon/WA Langsung ke Kampus", "Tanya langsung ke Humas / Bagian PMB / BAAK untuk data terbaru"),
]
for i, (src, desc) in enumerate(sources):
    r = 18 + i
    ws_cover.row_dimensions[r].height = 22
    bg = LIGHT_BG if i % 2 == 0 else WHITE

    c_src = ws_cover.cell(row=r, column=2, value=f"→  {src}")
    c_src.font = Font(bold=True, size=9, color=BLUE_FG, name="Arial")
    c_src.fill = fill(bg)
    c_src.alignment = align("left", "center")
    c_src.border = border()

    ws_cover.merge_cells(f"C{r}:H{r}")
    c_desc = ws_cover.cell(row=r, column=3, value=desc)
    c_desc.font = Font(size=9, color=GRAY_FG, name="Arial")
    c_desc.fill = fill(bg)
    c_desc.alignment = align("left", "center")
    c_desc.border = border()


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — CHECKLIST
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("✅ Checklist")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A3"

col_widths = {"A": 5, "B": 28, "C": 38, "D": 14, "E": 16, "F": 32, "G": 40}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# Title row
ws.row_dimensions[1].height = 36
ws.merge_cells("A1:G1")
c = ws["A1"]
c.value = "✅  CHECKLIST RISET DATA — KNOWLEDGE BASE UMBandung"
c.font = Font(bold=True, size=13, color=WHITE, name="Arial")
c.fill = fill(NAVY)
c.alignment = align("center", "center")

# Header row
headers = ["No", "Kategori & Fitur", "Deskripsi / Yang Perlu Dicari", "Prioritas", "Status", "Sumber Data", "Catatan / Data Ditemukan"]
ws.row_dimensions[2].height = 30
for i, h in enumerate(headers, 1):
    c = ws.cell(row=2, column=i, value=h)
    c.font = Font(bold=True, size=10, color=WHITE, name="Arial")
    c.fill = fill(BLUE)
    c.alignment = align("center", "center", wrap=True)
    c.border = border()

# Dropdown validation
dv = DataValidation(type="list",
                    formula1='"✅ Selesai,🔄 Proses,⏳ Belum,⛔ Skip"',
                    allow_blank=True)
dv.error = "Pilih dari dropdown"
dv.errorTitle = "Nilai tidak valid"
ws.add_data_validation(dv)

# Data rows
CATEGORIES = [
    # (category_label, bg_header_hex, items)
    ("A. IDENTITAS & PROFIL UNIVERSITAS", NAVY, [
        ("Profil Umum", "Nama lengkap & singkatan resmi universitas", "🔴 Penting"),
        ("Profil Umum", "Tahun berdiri & sejarah singkat universitas", "🟡 Normal"),
        ("Profil Umum", "Visi, Misi, dan Nilai Universitas", "🟡 Normal"),
        ("Profil Umum", 'Tagline resmi: "Islamic Technopreneurial University"', "🟡 Normal"),
        ("Profil Umum", "Status & nilai akreditasi institusi (BAN-PT), nomor SK, masa berlaku", "🔴 Penting"),
        ("Profil Umum", "Nama Rektor & pimpinan utama yang aktif saat ini", "🔴 Penting"),
        ("Profil Umum", "Afiliasi Muhammadiyah — PTM ke berapa, peran di lingkungan Muhammadiyah", "🟢 Opsional"),
        ("Lokasi & Kontak", "Alamat lengkap kampus (jalan, kelurahan, kecamatan, kota, kode pos)", "🔴 Penting"),
        ("Lokasi & Kontak", "Link Google Maps kampus utama", "🟡 Normal"),
        ("Lokasi & Kontak", "Nomor telepon & WhatsApp resmi kampus", "🔴 Penting"),
        ("Lokasi & Kontak", "Email resmi (info@, humas@, pmb@umbandung.ac.id)", "🟡 Normal"),
        ("Lokasi & Kontak", "Website & semua akun media sosial resmi", "🟡 Normal"),
        ("Lokasi & Kontak", "Akses transportasi umum: angkot, Trans Metro Bandung, titik turun terdekat", "🟢 Opsional"),
    ]),
    ("B. PENERIMAAN MAHASISWA BARU (PMB)", BLUE, [
        ("Jalur & Syarat", "Daftar semua jalur masuk yang tersedia (SNBT, Mandiri, Transfer, dll)", "🔴 Penting"),
        ("Jalur & Syarat", "Syarat umum pendaftaran (ijazah, nilai minimal, usia, dokumen)", "🔴 Penting"),
        ("Jalur & Syarat", "Jadwal PMB — tanggal buka & tutup tiap gelombang pendaftaran", "🟡 Normal"),
        ("Jalur & Syarat", "Biaya pendaftaran per jalur masuk", "🟡 Normal"),
        ("Jalur & Syarat", "Link pendaftaran online resmi (pmb.umbandung.ac.id atau portal lain)", "🔴 Penting"),
        ("Jalur & Syarat", "Alur tahapan seleksi (Daftar → Bayar → Tes → Pengumuman → Daftar Ulang)", "🟡 Normal"),
        ("Jalur & Syarat", "Nomor WhatsApp / email khusus panitia PMB", "🔴 Penting"),
        ("Biaya Kuliah", "UKT / biaya per semester untuk tiap program studi (kisaran)", "🔴 Penting"),
        ("Biaya Kuliah", "Sistem pembayaran: per SKS atau per semester", "🟡 Normal"),
        ("Biaya Kuliah", "Biaya DPP / uang pangkal jika ada", "🟡 Normal"),
        ("Biaya Kuliah", "Komponen biaya lain: seragam, almamater, KTM, orientasi", "🟡 Normal"),
        ("Biaya Kuliah", "Metode pembayaran yang diterima (bank transfer, virtual account, dll)", "🟢 Opsional"),
    ]),
    ("C. PROGRAM STUDI & FAKULTAS", "1A6FC4", [
        ("Fakultas & Prodi", "Daftar semua fakultas beserta nama resmi dan singkatannya", "🔴 Penting"),
        ("Fakultas & Prodi", "Daftar lengkap program studi S1 — nama prodi, jenjang, akreditasi tiap prodi", "🔴 Penting"),
        ("Fakultas & Prodi", "Program D3 / D4 jika tersedia", "🟡 Normal"),
        ("Fakultas & Prodi", "Program Pascasarjana S2 / S3 jika tersedia", "🟡 Normal"),
        ("Fakultas & Prodi", "Konsentrasi / peminatan dalam tiap program studi jika ada", "🟢 Opsional"),
        ("Fakultas & Prodi", "Prospek karir & profil lulusan per program studi", "🟡 Normal"),
    ]),
    ("D. BEASISWA & BANTUAN KEUANGAN", "3B6D11", [
        ("Beasiswa", "Beasiswa dari UMBandung sendiri — nama program, syarat, besaran, kuota", "🔴 Penting"),
        ("Beasiswa", "Beasiswa dari Persyarikatan Muhammadiyah / PP Muhammadiyah", "🟡 Normal"),
        ("Beasiswa", "KIP Kuliah — apakah UMBandung penerima? Kuota, cara daftar, syarat", "🔴 Penting"),
        ("Beasiswa", "Beasiswa prestasi akademik & non-akademik (juara lomba, hafidz, dll)", "🟡 Normal"),
        ("Beasiswa", "Beasiswa dari instansi / perusahaan mitra kampus jika ada", "🟢 Opsional"),
        ("Beasiswa", "Kontak / unit yang mengurus beasiswa (Bagian Kemahasiswaan / BAK)", "🟡 Normal"),
    ]),
    ("E. FASILITAS KAMPUS", "0F6E56", [
        ("Fasilitas", "Gedung & ruang yang tersedia — kuliah, lab, studio, seminar, aula", "🟡 Normal"),
        ("Fasilitas", "Perpustakaan — jam buka, koleksi, akses e-library / digital", "🟡 Normal"),
        ("Fasilitas", "Laboratorium & studio per prodi — nama lab, peralatan utama, jam akses", "🟡 Normal"),
        ("Fasilitas", "Fasilitas olahraga, kemahasiswaan, kantin, masjid", "🟡 Normal"),
        ("Fasilitas", "Koneksi internet / WiFi kampus — jangkauan, cara akses", "🟢 Opsional"),
        ("Fasilitas", "Asrama mahasiswa jika ada — kapasitas, lokasi, biaya, cara daftar", "🟢 Opsional"),
    ]),
    ("F. LAYANAN AKADEMIK & ADMINISTRASI", "534AB7", [
        ("Layanan Admin", "Cara mengurus KTM (baru, hilang, rusak) — prosedur dan biaya", "🟡 Normal"),
        ("Layanan Admin", "Cara mendapatkan surat keterangan aktif kuliah", "🔴 Penting"),
        ("Layanan Admin", "Prosedur transkrip nilai & legalisir ijazah — waktu proses & biaya", "🟡 Normal"),
        ("Layanan Admin", "Jam operasional Biro Akademik (BAAK) — hari, jam, kontak", "🟡 Normal"),
        ("Layanan Admin", "Portal akademik mahasiswa — nama sistem (SIAKAD/SIACER), link akses", "🟡 Normal"),
        ("Layanan Admin", "Prosedur pengajuan cuti kuliah dan aktif kembali", "🟢 Opsional"),
        ("Kalender Akademik", "Tanggal awal & akhir semester ganjil dan genap", "🟡 Normal"),
        ("Kalender Akademik", "Periode KRS / pengambilan mata kuliah", "🟡 Normal"),
        ("Kalender Akademik", "Jadwal UTS dan UAS", "🟡 Normal"),
        ("Kalender Akademik", "Jadwal libur nasional & libur akademik kampus", "🟢 Opsional"),
        ("Kalender Akademik", "Jadwal wisuda — periode, syarat, biaya pendaftaran wisuda", "🔴 Penting"),
    ]),
    ("G. KEMAHASISWAAN & KEGIATAN", "854F0B", [
        ("Organisasi & UKM", "BEM & Dewan Mahasiswa — struktur, kontak, kegiatan utama", "🟢 Opsional"),
        ("Organisasi & UKM", "Daftar UKM aktif — olahraga, seni, penelitian, kerohanian, dll", "🟡 Normal"),
        ("Organisasi & UKM", "Cara bergabung UKM & waktu pendaftaran anggota baru", "🟡 Normal"),
        ("Organisasi & UKM", "Organisasi Islam kampus: IMM, Tapak Suci, Hizbul Wathan", "🟡 Normal"),
        ("KKN & Magang", "Prosedur KKN — syarat SKS, cara daftar, lokasi penempatan", "🟡 Normal"),
        ("KKN & Magang", "Prosedur magang / PKL — difasilitasi kampus atau mandiri", "🟡 Normal"),
        ("KKN & Magang", "MBKM — program yang tersedia (pertukaran pelajar, proyek desa, dll)", "🟡 Normal"),
        ("KKN & Magang", "Pusat penelitian / lembaga riset yang ada di UMBandung", "🟢 Opsional"),
    ]),
    ("H. KELULUSAN & ALUMNI", "A32D2D", [
        ("Tugas Akhir & Wisuda", "Syarat mengajukan skripsi / tugas akhir — minimal SKS, nilai, cara daftar", "🔴 Penting"),
        ("Tugas Akhir & Wisuda", "Syarat pendaftaran wisuda — bebas tanggungan, bebas perpustakaan, biaya", "🔴 Penting"),
        ("Tugas Akhir & Wisuda", "Predikat kelulusan dan batas IPK (Cumlaude, Sangat Memuaskan, dll)", "🟡 Normal"),
        ("Alumni", "Ikatan Alumni UMBandung (IKA UMB) — cara bergabung, manfaat, kontak", "🟢 Opsional"),
    ]),
]

row = 3
item_num = 1
for cat_label, cat_color, items in CATEGORIES:
    # Category header row
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value=f"  {cat_label}")
    c.font = Font(bold=True, size=10, color=WHITE, name="Arial")
    c.fill = fill(cat_color)
    c.alignment = align("left", "center")
    c.border = border()
    row += 1

    for (sub, desc, priority) in items:
        ws.row_dimensions[row].height = 32
        bg = WHITE if item_num % 2 == 0 else ROW_ALT

        # Col A — number
        c = ws.cell(row=row, column=1, value=item_num)
        c.font = font(size=9, color=GRAY_FG)
        c.fill = fill(bg)
        c.alignment = align("center", "center")
        c.border = border()

        # Col B — subcategory + feature
        c = ws.cell(row=row, column=2, value=f"[{sub}]\n{desc.split(' — ')[0]}")
        c.font = Font(bold=True, size=9, color=NAVY, name="Arial")
        c.fill = fill(bg)
        c.alignment = align("left", "center", wrap=True)
        c.border = border()

        # Col C — description
        c = ws.cell(row=row, column=3, value=desc)
        c.font = font(size=9, color="374151")
        c.fill = fill(bg)
        c.alignment = align("left", "center", wrap=True)
        c.border = border()

        # Col D — priority
        prio_bg = RED_BG if "Penting" in priority else (AMBER_BG if "Normal" in priority else GREEN_BG)
        prio_fg = RED_FG if "Penting" in priority else (AMBER_FG if "Normal" in priority else GREEN_FG)
        c = ws.cell(row=row, column=4, value=priority)
        c.font = Font(bold=True, size=9, color=prio_fg, name="Arial")
        c.fill = fill(prio_bg)
        c.alignment = align("center", "center")
        c.border = border()

        # Col E — status (dropdown)
        c = ws.cell(row=row, column=5, value="⏳ Belum")
        c.font = font(size=9, color=AMBER_FG)
        c.fill = fill(AMBER_BG)
        c.alignment = align("center", "center")
        c.border = border()
        dv.add(c)

        # Col F — source
        c = ws.cell(row=row, column=6, value="")
        c.font = font(size=9, color=BLUE_FG)
        c.fill = fill(bg)
        c.alignment = align("left", "center", wrap=True)
        c.border = border()

        # Col G — notes
        c = ws.cell(row=row, column=7, value="")
        c.font = font(size=9, color="374151")
        c.fill = fill(bg)
        c.alignment = align("left", "center", wrap=True)
        c.border = border()

        item_num += 1
        row += 1


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — TEMPLATE KNOWLEDGE BASE
# ══════════════════════════════════════════════════════════════════════════════
ws_kb = wb.create_sheet("📝 Template KB")
ws_kb.sheet_view.showGridLines = False
ws_kb.freeze_panes = "A3"

ws_kb.column_dimensions["A"].width = 5
ws_kb.column_dimensions["B"].width = 36
ws_kb.column_dimensions["C"].width = 14
ws_kb.column_dimensions["D"].width = 60
ws_kb.column_dimensions["E"].width = 20

ws_kb.row_dimensions[1].height = 36
ws_kb.merge_cells("A1:E1")
c = ws_kb["A1"]
c.value = "📝  TEMPLATE KNOWLEDGE BASE — Siap Dimasukkan ke database_ID.py"
c.font = Font(bold=True, size=13, color=WHITE, name="Arial")
c.fill = fill(NAVY)
c.alignment = align("center", "center")

kb_headers = ["No", "Keyword / Key (Pertanyaan Trigger)", "Tipe Respons", "Konten Jawaban (HTML / Teks)", "Catatan Tambahan"]
ws_kb.row_dimensions[2].height = 28
for i, h in enumerate(kb_headers, 1):
    c = ws_kb.cell(row=2, column=i, value=h)
    c.font = Font(bold=True, size=10, color=WHITE, name="Arial")
    c.fill = fill(BLUE)
    c.alignment = align("center", "center", wrap=True)
    c.border = border()

dv_type = DataValidation(type="list", formula1='"paragraph,list,table"', allow_blank=True)
ws_kb.add_data_validation(dv_type)

kb_examples = [
    ("profil umbandung", "paragraph",
     "Universitas Muhammadiyah Bandung (UMBandung) adalah perguruan tinggi swasta berbasis Islam...",
     "Isi dari about/profil di website"),
    ("alamat kampus umbandung", "paragraph",
     "Kampus UMBandung berlokasi di Jl. [ISIAN]. Koordinat GPS: [...]. <br><a href='[LINK MAPS]'>Buka di Google Maps</a>",
     "Tambahkan link Google Maps aktif"),
    ("cara daftar pmb umbandung", "paragraph",
     "<strong>Langkah Pendaftaran PMB UMBandung</strong><ol><li>Kunjungi [LINK PMB]</li><li>Buat akun & isi formulir</li>...</ol>",
     "Update tiap tahun ajaran baru"),
    ("biaya kuliah umbandung", "paragraph",
     "<strong>Informasi Biaya Kuliah</strong><br>UKT UMBandung berkisar antara Rp [X] - Rp [Y] per semester tergantung program studi...",
     "Cantumkan rentang, bukan nominal pasti"),
    ("program studi umbandung", "paragraph",
     "<strong>Program Studi UMBandung</strong><ul><li>Fakultas [...]: Prodi A, Prodi B</li><li>Fakultas [...]: Prodi C</li></ul>",
     "Update jika ada prodi baru"),
    ("beasiswa umbandung", "paragraph",
     "<strong>Program Beasiswa UMBandung</strong><ul><li>Beasiswa Prestasi: [syarat]</li><li>KIP Kuliah: [syarat]</li></ul>",
     "Cek bagian kemahasiswaan"),
    ("akreditasi umbandung", "paragraph",
     "Universitas Muhammadiyah Bandung terakreditasi [NILAI] oleh BAN-PT berdasarkan SK No. [...] berlaku hingga [TAHUN].",
     "Update setiap perpanjangan akreditasi"),
    ("kontak umbandung", "paragraph",
     "<strong>Kontak UMBandung</strong><br>Telepon: [NO]<br>WhatsApp: [NO]<br>Email: [EMAIL]<br>Website: umbandung.ac.id",
     "Tambahkan jam operasional"),
    ("jadwal wisuda umbandung", "paragraph",
     "<strong>Wisuda UMBandung</strong><br>Wisuda diselenggarakan [X] kali per tahun. Syarat pendaftaran: [ISIAN].",
     "Update tiap periode"),
    ("syarat skripsi umbandung", "paragraph",
     "<strong>Syarat Pengajuan Skripsi/TA</strong><ul><li>Minimal [X] SKS lulus</li><li>IPK minimal [X]</li><li>Sudah lulus mata kuliah wajib</li></ul>",
     "Cek BAAK atau panduan akademik"),
]

for i, (key, rtype, content, note) in enumerate(kb_examples, 1):
    r = i + 2
    ws_kb.row_dimensions[r].height = 40
    bg = WHITE if i % 2 == 0 else ROW_ALT

    c = ws_kb.cell(row=r, column=1, value=i)
    c.font = font(size=9, color=GRAY_FG); c.fill = fill(bg)
    c.alignment = align("center","center"); c.border = border()

    c = ws_kb.cell(row=r, column=2, value=key)
    c.font = Font(bold=True, size=9, color=BLUE_FG, name="Arial"); c.fill = fill(bg)
    c.alignment = align("left","center",wrap=True); c.border = border()

    c = ws_kb.cell(row=r, column=3, value=rtype)
    c.font = font(size=9, color=PURPLE_FG); c.fill = fill(PURPLE_BG)
    c.alignment = align("center","center"); c.border = border()
    dv_type.add(c)

    c = ws_kb.cell(row=r, column=4, value=content)
    c.font = font(size=9, color="374151"); c.fill = fill(bg)
    c.alignment = align("left","center",wrap=True); c.border = border()

    c = ws_kb.cell(row=r, column=5, value=note)
    c.font = Font(italic=True, size=9, color=GRAY_FG, name="Arial"); c.fill = fill(bg)
    c.alignment = align("left","center",wrap=True); c.border = border()

# ── Add legend row at bottom ──
r_legend = len(kb_examples) + 4
ws_kb.row_dimensions[r_legend].height = 20
ws_kb.merge_cells(f"A{r_legend}:E{r_legend}")
c = ws_kb.cell(row=r_legend, column=1,
    value="💡  Ganti semua teks dalam [KURUNG SIKU] dengan data aktual UMBandung hasil riset kamu")
c.font = Font(italic=True, size=9, color=AMBER_FG, name="Arial")
c.fill = fill(AMBER_BG)
c.alignment = align("left","center")
c.border = border()


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — AUTOCOMPLETE TERMS
# ══════════════════════════════════════════════════════════════════════════════
ws_ac = wb.create_sheet("🔤 Autocomplete Terms")
ws_ac.sheet_view.showGridLines = False
ws_ac.column_dimensions["A"].width = 5
ws_ac.column_dimensions["B"].width = 42
ws_ac.column_dimensions["C"].width = 22
ws_ac.column_dimensions["D"].width = 32

ws_ac.row_dimensions[1].height = 36
ws_ac.merge_cells("A1:D1")
c = ws_ac["A1"]
c.value = "🔤  AUTOCOMPLETE TERMS — Untuk AUTOCOMPLETE_TERMS_ID di database_ID.py"
c.font = Font(bold=True, size=13, color=WHITE, name="Arial")
c.fill = fill(NAVY)
c.alignment = align("center","center")

ac_headers = ["No", "Term (Kata Kunci Autocomplete)", "Kategori", "Keterangan"]
ws_ac.row_dimensions[2].height = 26
for i, h in enumerate(ac_headers, 1):
    c = ws_ac.cell(row=2, column=i, value=h)
    c.font = Font(bold=True, size=10, color=WHITE, name="Arial")
    c.fill = fill(BLUE); c.alignment = align("center","center",wrap=True)
    c.border = border()

ac_terms = [
    ("cara daftar PMB Universitas Muhammadiyah Bandung", "PMB", "Paling sering dicari calon mhs"),
    ("biaya kuliah UMBandung 2025", "PMB/Biaya", "Update tahun setiap tahun ajaran"),
    ("jalur masuk UMBandung", "PMB", ""),
    ("syarat pendaftaran UMBandung", "PMB", ""),
    ("jadwal PMB UMBandung", "PMB", ""),
    ("beasiswa UMBandung", "Beasiswa", ""),
    ("KIP Kuliah UMBandung", "Beasiswa", ""),
    ("program studi UMBandung", "Prodi", ""),
    ("fakultas UMBandung", "Prodi", ""),
    ("akreditasi UMBandung", "Profil", ""),
    ("alamat kampus UMBandung", "Kontak", ""),
    ("nomor telepon UMBandung", "Kontak", ""),
    ("WhatsApp UMBandung", "Kontak", ""),
    ("email UMBandung", "Kontak", ""),
    ("lokasi kampus UMBandung", "Kontak", ""),
    ("jadwal wisuda UMBandung", "Akademik", ""),
    ("syarat wisuda UMBandung", "Akademik", ""),
    ("syarat skripsi UMBandung", "Akademik", ""),
    ("KRS UMBandung", "Akademik", ""),
    ("portal akademik UMBandung", "Akademik", ""),
    ("BAAK UMBandung", "Layanan", ""),
    ("surat keterangan aktif UMBandung", "Layanan", ""),
    ("transkrip nilai UMBandung", "Layanan", ""),
    ("UKM UMBandung", "Kemahasiswaan", ""),
    ("BEM UMBandung", "Kemahasiswaan", ""),
    ("KKN UMBandung", "Kemahasiswaan", ""),
    ("magang UMBandung", "Kemahasiswaan", ""),
    ("MBKM UMBandung", "Kemahasiswaan", ""),
    ("perpustakaan UMBandung", "Fasilitas", ""),
    ("laboratorium UMBandung", "Fasilitas", ""),
    ("asrama UMBandung", "Fasilitas", ""),
    ("wifi kampus UMBandung", "Fasilitas", ""),
    ("visi misi UMBandung", "Profil", ""),
    ("sejarah UMBandung", "Profil", ""),
    ("rektor UMBandung", "Profil", ""),
]

for i, (term, cat, note) in enumerate(ac_terms, 1):
    r = i + 2
    ws_ac.row_dimensions[r].height = 22
    bg = WHITE if i % 2 == 0 else ROW_ALT

    c = ws_ac.cell(row=r, column=1, value=i)
    c.font = font(size=9, color=GRAY_FG); c.fill = fill(bg)
    c.alignment = align("center","center"); c.border = border()

    c = ws_ac.cell(row=r, column=2, value=term)
    c.font = font(size=9); c.fill = fill(bg)
    c.alignment = align("left","center"); c.border = border()

    c = ws_ac.cell(row=r, column=3, value=cat)
    c.font = Font(bold=True, size=9, color=BLUE_FG, name="Arial"); c.fill = fill(BLUE_BG)
    c.alignment = align("center","center"); c.border = border()

    c = ws_ac.cell(row=r, column=4, value=note)
    c.font = Font(italic=True, size=9, color=GRAY_FG, name="Arial"); c.fill = fill(bg)
    c.alignment = align("left","center"); c.border = border()


# ── Tab colors ─────────────────────────────────────────────────────────────────
ws_cover.sheet_properties.tabColor  = NAVY
ws.sheet_properties.tabColor         = "0D4EA6"
ws_kb.sheet_properties.tabColor      = "3B6D11"
ws_ac.sheet_properties.tabColor      = "534AB7"

# ── Print settings ─────────────────────────────────────────────────────────────
for sheet in [ws_cover, ws, ws_kb, ws_ac]:
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToPage = True
    sheet.page_setup.fitToWidth = 1

wb.save("Riset_Data_UMBandung.xlsx")
print("DONE")